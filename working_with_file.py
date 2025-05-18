import csv
import os
import io
from zipfile import ZipFile
import pytest
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader

PATH_RES ="resources"
PATH_ARCH = "my_zip.zip"

L_files = os.listdir(PATH_RES) #создаем список для удобства

def create_archive(): #создаем архив с файлами
    with ZipFile(PATH_ARCH, mode="w") as myzip:
        for file in L_files:
            myzip.write(os.path.join(PATH_RES, file)) #файлы вынулись из директории и из них создался архив
#create_archive()

@pytest.fixture(scope="session")
def myzip():
    with ZipFile(PATH_ARCH, "r") as myzip:
        yield myzip

@pytest.fixture
def archived_file(myzip, request):
    file_format = request.param
    for file in myzip.namelist():
        if file.endswith(file_format):
            with myzip.open(file, "r") as fd:
                yield fd

@pytest.mark.parametrize('archived_file', ["csv"], indirect=True)
def test_csv(archived_file):
    with io.TextIOWrapper(archived_file, encoding="utf-8") as io_wrapper:
        csvfile = csv.reader(io_wrapper)
        founded = False
        for row in csvfile:
            if row == ['3', 'Sarena', 'Cornewell', 'scornewell2@nyu.edu', 'Female',
                       '251.107.184.114']:  # проверка наличия строки
                founded = True
                break
        assert founded == True

@pytest.mark.parametrize('archived_file', ["xlsx"], indirect=True)
def test_xlsx(archived_file):
    xlsxfile = load_workbook(archived_file)
    sheet = xlsxfile.active
    assert sheet.cell(row=1, column=1).value == 'snail!'

@pytest.mark.parametrize('archived_file', ["pdf"], indirect=True)
def test_pdf(archived_file):
    pdf_file = PdfReader(archived_file)
    page = pdf_file.pages[0]
    text = page.extract_text()
    assert "Никакой полезной информации он не несёт" in text
